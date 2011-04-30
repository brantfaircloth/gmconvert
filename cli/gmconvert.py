#!/usr/bin/env python
# encoding: utf-8

"""
gmconvert.py

Created by Brant Faircloth on 28 April 2011.
Copyright 2011 Brant C. Faircloth. All rights reserved.
"""


import pdb
import os
import sys
import copy
import optparse
import numpy
import la
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.writer.excel import ExcelWriter



def interface():
    '''Command-line interface'''
    usage = "usage: %prog [options]"

    p = optparse.OptionParser(usage)

    p.add_option('--input', dest = 'input', action='store', 
type='string', default = None, help='The path to the input file.', 
metavar='FILE')

    p.add_option('--output', dest = 'output', action='store', 
type='string', default = None, help='The path to the output file.', 
metavar='FILE')

    (options,arg) = p.parse_args()
    if not options.input:
        p.print_help()
        sys.exit(2)
    if not os.path.isfile(options.input):
        print "You must provide a valid path to the configuration file."
        p.print_help()
        sys.exit(2)
    return options, arg

def get_number_of_markers(input):
    markers = set()
    samples = set()
    for line in open(input, 'rU'):
        line_elements = line.strip('\n').split('\t')
        if line_elements[0] == 'Sample File':
            marker_pos = line_elements.index('Marker')
            sample_pos = line_elements.index('Sample Name')
        else:
            markers.add(line_elements[marker_pos])
            samples.add(line_elements[sample_pos])
    return markers, samples
    
def create_population(samples, markers, columns = ['A','B']):
    population = {}
    for individual in samples:
        population[individual] = create_larry(markers)
    return population

def create_larry(markers, depth = '0', columns = ['A','B']):
    labels = [[depth], list(markers), columns]
    a = numpy.zeros((1,len(markers),len(columns)), dtype='|S4')
    return la.larry(a , labels)
    
def insert_genotype_data(input, population, markers):
    for line in open(input, 'rU'):
        line_elements = line.strip('\n').split('\t')
        if line_elements[0] == 'Sample File':
            sample_pos = line_elements.index('Sample Name')
            marker_pos = line_elements.index('Marker')
            allele1_pos = line_elements.index('Allele 1')
            allele2_pos = line_elements.index('Allele 2')
        else:
            name = line_elements[sample_pos]
            marker = line_elements[marker_pos]
            a1 = line_elements[allele1_pos]
            a2 = line_elements[allele2_pos]
            if a1 == '' and a2 == '':
                pass
            else:
                if a2 == '' and a1 != '':
                    a2 = a1
                # check for entries in first level of larry
                original_shape = population[name].shape[0]
                for l in xrange(original_shape):
                    level = str(l)
                    #pdb.set_trace()
                    if (population[name].lix[[level]].lix[[marker]].x == numpy.array(['', ''], dtype='|S4')).all():
                        population[name].lix[[level]].lix[[marker]].x[0] = a1
                        population[name].lix[[level]].lix[[marker]].x[1] = a2
                    elif not original_shape > (l + 1):
                        #pdb.set_trace()
                        new_level = "{0}".format(l + 1)
                        # create a new level
                        empty = create_larry(markers, depth = new_level)
                        # merge it into the existing level
                        population[name] = population[name].merge(empty)
                        population[name].lix[[new_level]].lix[[marker]].x[0] = a1
                        population[name].lix[[new_level]].lix[[marker]].x[1] = a2
    return population

def write_header(ws, header):
    # excel columns are indexed by 1, not 0
    for col in xrange(1, len(header) + 1):
        cl = get_column_letter(col)
        ws.cell('{0}1'.format(cl)).value = header[col - 1]
    return ws

def get_max_depth(population):
    depths = []
    for name, genotypes in population.iteritems():
        depths.append(genotypes.shape[0])
    return max(depths)

def create_additional_sheet(wb, header, max_depth):
    sheet_depth = len(wb.get_sheet_names())
    if sheet_depth < max_depth:
        for sheet in range(max_depth - sheet_depth):
            ws = wb.create_sheet()
            ws.title = "Level{0}".format(sheet + sheet_depth)
            ws = write_header(ws, header)
    # create a summary sheet
    ws = wb.create_sheet()
    ws.title = "Summary"
    ws = write_header(ws, header)
    return wb
    
def write_records_to_excel(output, population, markers, alleles = ['A','B']):
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    ws = wb.worksheets[0]
    ws.title = "Level0"
    # add marker columns - 2 columns for each allele
    header = ["Individual"] + ["{0}_{1}".format(m,a) for m in markers for a in alleles]
    #pdb.set_trace()
    ws = write_header(ws, header)
    max_depth = get_max_depth(population)
    wb = create_additional_sheet(wb, header, max_depth)
    pop_keys = population.keys()
    pop_keys.sort()
    for name_idx, name in enumerate(pop_keys):
        #   pdb.set_trace()
        # get depth of stack
        depth = population[name].shape[0]
        # write invididual name at all depths
        #for level in xrange(max_depth):
        #    ws = wb.get_sheet_by_name("Level{0}".format(level))
        #    ws.cell('A{0}'.format(name_idx + 2)).value = name
        # ensure there is workbook at max stack depth
        for level in xrange(max_depth):
            # write the sample id for each row of all levels
            ws = wb.get_sheet_by_name("Level{0}".format(level))
            ws.cell('A{0}'.format(name_idx + 2)).set_value_explicit(value = name, data_type = 's')
            # but only write the genotypes for the sample where there 
            # is a level
            if level < depth:
                for marker_idx, marker in enumerate(header[1:]):
                    cl = get_column_letter(marker_idx + 2)
                    marker, allele = marker.split('_')
                    if allele == 'A':
                        pos = 0
                    else:
                        pos = 1
                    ws.cell('{0}{1}'.format(cl, name_idx + 2)).value = population[name].lix[[str(level)]].lix[[marker]].x[pos]
        # check all non-zero entries for similarity and store to summary
        ws = wb.get_sheet_by_name("Summary".format(level))
        ws.cell('A{0}'.format(name_idx + 2)).set_value_explicit(value = name, data_type = 's')
        for marker_idx, marker in enumerate(header[1:]):
            cl = get_column_letter(marker_idx + 2)
            marker, allele = marker.split('_')
            if allele == 'A':
                pos = 0
            else:
                pos = 1
            # if only one element, self = self
            if depth <= 1:
                identical = True
            elif (population[name].lix[:,[marker],pos].x == '').all():
                identical = True
            else:
                # don't penalize comparisons having '' in >= one column
                empties = population[name].lix[:,[marker],pos].x != ''
                genotypes_no_empties = population[name].lix[:,[marker],pos].x[empties]
                # just test first element against all elements
                identical = (genotypes_no_empties == genotypes_no_empties[0]).all()
            if identical:
                ws.cell('{0}{1}'.format(cl, name_idx + 2)).value = "T"
            else:
                ws.cell('{0}{1}'.format(cl, name_idx + 2)).value = "F"
        #pdb.set_trace()
    ew.save(filename = output)

def main():
    options, args = interface()
    markers, samples = get_number_of_markers(options.input)
    m_list = list(markers)
    population = create_population(samples, markers)
    population = insert_genotype_data(options.input, population, markers)
    write_records_to_excel(options.output, population, markers)
    #pdb.set_trace()
        
if __name__ == '__main__':
    main()